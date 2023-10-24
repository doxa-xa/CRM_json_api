from flask import Flask, request, url_for, flash, render_template, Response, send_from_directory, redirect
from flask.logging import default_handler
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
from models import Customer, Product, db
from customer_response import CustomerResponse
from product import ProductResponse
from utils import strToBool
import logging
import datetime as dt
import os
import openpyxl
import csv
import datetime

#TODO: REDUCE THE CODE REPETITION IN HELPER METHODS IMPROVE SPEED

UPLOAD_FOLDER = './uploaded'
ALLOWED_EXTENTIONS = {'xls','xlsx'}

app = Flask(__name__)
app.secret_key = 'this is super secret key'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///customers.db'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
ALLOWED_EXTENSIONS = {'xlsx','xls','txt','csv'}

db.init_app(app)

with app.app_context():
    db.create_all()

def allowed_file(filename):
    return '.' in filename and \
        filename.rsplit('.',1)[1].lower() in ALLOWED_EXTENSIONS

def db_upload_customers(file):
    filetype = file.rsplit('.',1)
    if filetype[1] == 'xls' or filetype[1] == 'xlsx':
        workbook = openpyxl.load_workbook(file)
        #excel workbook has to have one sheet, it will only select the first one, thus [0]
        ws_name = workbook.get_sheet_names()[0]
        worksheet = workbook.get_sheet_by_name(ws_name)  # -> TODO: Update the method or create one for several sheets        
        for row in worksheet.iter_rows():
            #print(row)
            if row[0] and row [1]: # -> TODO: Update verificaton for better & accurate upload
                customer = Customer(
                    name=row[0].value,
                    email=row[1].value,
                    address=row[2].value,
                    phone=row[3].value,
                    opt_email=row[4].value,
                    opt_phone=row[5].value,
                    opt_chat=row[6].value,
                    status=row[7].value,
                    last_updated=datetime.datetime.now()
                    #last_contacted=row[8].value -> TODO: helper method for parsing date from excel
                    )
                db.session.add(customer)
                db.session.commit()
            else:
                logging.waring(f'customer not uploaded in db, missing data name and email: {row}')
    elif filetype == 'csv':
        with open(file,'r', encoding='utf-8') as csv_file:
            reader = csv.reader(csv_file)
            for row in reader:
                if row[0] and row[1]:
                    customer = Customer(
                        name=row[0],
                        email=row[1],
                        address=row[2],
                        phone=row[3],
                        opt_email=row[4],
                        opt_phone=row[5],
                        opt_chat=row[6],
                        status=row[7],
                        last_updated=datetime.datetime.now()
                    )
                    db.session.add(customer)
    elif filetype == 'txt':
        pass # -> TODO to implement txt files read
    db.session.commit()

def db_upload_products(file):
    filetype = file.rsplit('.',1)
    if filetype[1] == 'xls' or filetype[1] == 'xlsx':
        workbook = openpyxl.load_workbook(file)
        #excel workbook has to have one sheet, it will only select the first one, thus [0]
        ws_name = workbook.get_sheet_names()[0]
        worksheet = workbook.get_sheet_by_name(ws_name)  # -> TODO: Update the method or create one for several sheets        
        for row in worksheet.iter_rows():
            #print(row)
            if row[0] and row [1]: # -> TODO: Update verificaton for better & accurate upload
                product = Product(
                    name=row[0].value,
                    price=float(row[1].value)
                    #purhcased=row[2].value, # -> TODO: helper method for parsing date from excel
                    #warranty=row[3].value,
                    )
                db.session.add(customer)
                db.session.commit()
            else:
                logging.waring(f'customer not uploaded in db, missing data name and email: {row}')
    elif filetype == 'csv':
        with open(file,'r', encoding='utf-8') as csv_file:
            reader = csv.reader(csv_file)
            for row in reader:
                if row[0] and row[1]:
                    customer = Customer(
                        name=row[0],
                        price=float(row[1])
                        #purchased=row[2], # -> TODO: helper method for parsing date from csv
                        #warranty=row[3],
                    )
                    db.session.add(customer)
    elif filetype == 'txt':
        pass # -> TODO to implement txt files read
    db.session.commit()


# GET routes -------------------------------------------

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/add')
def add():
    return render_template('addproduct.html')

#Query customer by id
@app.route('/customerid=<int:id>')
def get_customer(id):
    cust = db.get_or_404(Customer,id)
    if cust:
        customer = CustomerResponse(cust)
        app.logger.info("%s cutomer found")
        return customer.json_response(), 200
    else:
        app.logger.warning("%s customer not found")
        return "Customer Not Found", 404
    
#Query product by id
@app.route('/productid=<int:id>')
def get_product(id):
    item = db.get_or_404(Product,id)
    if item:
        product = ProductResponse(item)
        app.logger.info("%s customer found")
        return product.json_response(), 200
    else:
        app.loger.info("%s customer not found")
        return "Customer Not Found", 404

#POST routes ------------------------------------------------

#Adding product
@app.route('/add/product',methods=['POST'])
def add_product():
    id = request.form.get('customerid')
    name = request.form.get('name')
    price = request.form.get('price')
    purchased = request.form.get('purchased')
    print(purchased)
    wrnty = request.form.get('warranty')
    date = purchased.split('-')
    yyyy = int(date[0])+int(wrnty)
    mm = int(date[1])
    dd = int(date[2])
    warranty = dt.date(yyyy,mm,dd)
    product = Product(
        name = name,
        price = float(price),
        purchased = dt.datetime.strptime(purchased, "%Y-%m-%d"),
        customer_id = int(id),
        warranty = warranty
    )
    with app.app_context():
        db.session.add(product)
        db.session.commit()
    app.logging.info(f"Product: {name} added to customer id: {id}")
    return f'Product: {name} added to customer id: {id}', 200

#adding customer
@app.route('/add/customer',methods=['POST'])
def add_customer():
    name = request.form.get('name')
    address = request.form.get('address')
    phone = request.form.get('phone')
    email = request.form.get('email')
    status = request.form.get('status')
    updated = dt.datetime.now()
    #contracted = request.form.get('contracted')
    opt_phone = request.form.get('optphone')
    opt_email = request.form.get('optemail')
    opt_chat = request.form.get('optchat')
    customer = Customer(name=name,
                        address=address,
                        phone=phone,
                        email=email,
                        status=status,
                        opt_chat=strToBool(opt_chat),
                        opt_phone=strToBool(opt_phone),
                        opt_email=strToBool(opt_email),
                        last_updated=updated,
                       )
    with app.app_context():
        db.session.add(customer)
        db.session.commit()
    app.logging.info(f"Customer: {name} added")
    return "customer added", 200

@app.route('/upload/customers',methods=['POST'])
def import_customers():
    if 'custfile' not in request.files:
        flash(f'No or corrupted file')
        return redirect(url_for('index'))
    file = request.files['custfile']
    if file.filename == '':
        flash(f'No file selected for upload {file}')
        return redirect(url_for('index'))
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        db_upload_customers(filepath)
        flash('File uploaded successfully')
        return redirect(url_for('index'))
    
@app.route('/upload/products',methods=['POST'])
def import_products():
    if 'prodfile' not in request.files:
        flash(f'No or corrupted file')
        return redirect(url_for('index'))
    file = request.files['prodfile']
    if file.filename == '':
        flash(f'No file selected for upload {file}')
        return redirect(url_for('index'))
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        db_upload_products(filepath)
        flash('File uploaded successfully')
        return redirect(url_for('index'))



    
logging.basicConfig(filename='./api.log',level=logging.INFO)
import api_routes_customer
import api_routes_product 
import api_routes_feedback
import api_routes_customer_requests

if __name__ == '__main__':
    app.run(debug=True)

